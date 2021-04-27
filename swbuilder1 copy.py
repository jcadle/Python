from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import random
import time
workbook = load_workbook(r"C:\Users\CadleJ\PycharmProjects\practice\Regis_EGS_list1.xlsx")
sheet = workbook.active
options = Options()
options.binary_location = r"C:\Users\CadleJ\AppData\Local\Mozilla Firefox\firefox.exe"

institutionURL = "https://works.bepress.com/institution/manage/fcc9ac87-9649-4654-8392-f44f368bfa9a/"
username = "jcadle@bepress.com"
password = "g3BUstaJ"
for row in sheet.iter_rows(min_row=2, values_only=True):
    person = {
        "email": row[0],
        "prefix": row[1],
        "first": row[2],
        "middle": row[3],
        "last": row[4],
        "suffix": row[5],
        "discipline": row[6],
        "ptype": row[7],
        "ptitle": row[8],
    }
    browser = webdriver.Firefox(options=options, executable_path="C:\sel\geckodriver.exe")
    browser.get(institutionURL)
    browser.find_element_by_name("username").send_keys(username)
    browser.find_element_by_name("password").send_keys(password)
    browser.find_element_by_xpath("//button[@type='submit']").click()
    browser.find_element_by_xpath("//button[@title='Create New Profile']").click()
    browser.find_element_by_id("id_add_email").send_keys(person["email"])
    browser.find_element_by_id("id_add_email").send_keys(Keys.TAB)
    time.sleep(5)
    browser.find_element_by_id("id_add_prefix").send_keys(person["prefix"])
    browser.find_element_by_id("id_add_first").send_keys(person["first"])
    browser.find_element_by_id("id_add_middle").send_keys(person["middle"])
    browser.find_element_by_id("id_add_last").send_keys(person["last"])
    browser.find_element_by_id("id_add_suffix").send_keys(person["suffix"])
    browser.find_element_by_link_text("terms of service").send_keys(Keys.TAB, Keys.ENTER)
    time.sleep(5)
    try:
        slugerror = str(browser.find_element_by_css_selector("span.xt-validation-item"))
    except:
        slugerror = ""
    generateslug = len(slugerror)
    if (generateslug > 0):
            slugnum = random.randint(10000,99999)
            browser.find_element_by_id("profilename").send_keys(slugnum)
    browser.find_element_by_name("search").send_keys(person["discipline"])
    discpath = "//span[text()='" + person["discipline"] + "']/preceding-sibling::span[@class='fancytree-checkbox']"
    browser.find_element_by_xpath(discpath).click()
    Select(browser.find_element_by_id("positon-add")).select_by_visible_text(person["ptype"])
    time.sleep(3)
    browser.find_element_by_id("new-experience-title").send_keys(person["ptitle"], Keys.TAB, Keys.TAB, Keys.TAB, Keys.ENTER)
    time.sleep(30)
    browser.close()
